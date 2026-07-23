"""
CS Weekly Performance Report - per-agent email (weekly, IST)  [v2 matrix layout]

Each rostered L1 / L2 agent gets a personal weekly email: (1) Team (their tier), (2) Shift
(their roster group), (3) You, each as a metrics x week matrix (latest week leftmost, small
good/bad arrow on the latest column only); a reopen-snapshot bucket x week matrix + an attached
per-agent reopen dump (.xlsx); and (4) AI notes written by an LLM from the numbers.

WHERE THE LOGIC LIVES - all numbers come from Redash; the DAG only slices, renders, and sends.
  * CORE_QUERY_ID   41791  - Team/Shift/Agent x week wide payload (closed-week, assignee-credited)
  * REOPEN_QUERY_ID 41792  - per reopen-event rows (last-closer attributed) -> buckets + xlsx + issue feed
  * CONFIG_QUERY_ID 41839  - from/reply-to address, alert channel, n_weeks, llm_* (OpenAI-compatible)
Config-in-Redash: edit the config row to change address / model / weeks with no code push.

DELIVERY: Gmail send is intentionally STUBBED (send_email raises unless creds wired). Run with
CS_PERF_DRY_RUN=1 to assemble from real data and write each agent's HTML + xlsx to disk (no send).

Schedule: Monday 19:00 Asia/Kolkata, after the prior Mon-Sun week has closed. Paused on creation.
"""

from datetime import timedelta, date
import logging, os, io, json, urllib.request

import pendulum
from airflow import DAG
from airflow.operators.python import PythonOperator

from utils.slack import RedashClient
from utils.slack.slack_config import REDASH_API_KEY, REDASH_BASE_URL

logger = logging.getLogger(__name__)

CORE_QUERY_ID     = 41791
REOPEN_QUERY_ID   = 41792
CONFIG_QUERY_ID   = 41839
BASELINES_QUERY_ID= 41854
DRY_RUN         = os.getenv('CS_PERF_DRY_RUN') == '1'
DRY_RUN_DIR     = os.getenv('CS_PERF_DRY_RUN_DIR', '/tmp/cs_perf_out')

# palette / ink
INK='#0f172a'; MUT='#64748b'; FAINT='#94a3b8'; LINE='#eef1f5'; ROW='#f4f6f9'
HL='#eef5fd'; HLB='#d7e6fb'; UP='#15803d'; DOWN='#b91c1c'; GREY='#b3b8c0'
BUCKET_ORDER=['incorrect','incomplete','new_issue','clarification','noise']
BUCKET_NICE={'incorrect':'Incorrect','incomplete':'Incomplete','new_issue':'New issue','clarification':'Clarification','noise':'Noise'}
AVOID_BUCKETS={'incorrect','incomplete'}

# ==================== FETCH ====================
def _redash():
    return RedashClient(api_key=REDASH_API_KEY, base_url=REDASH_BASE_URL)

def fetch_all():
    r=_redash()
    cfg=r.fetch_query_results(query_id=CONFIG_QUERY_ID, max_retries=3)[0]
    core=r.fetch_query_results(query_id=CORE_QUERY_ID, max_retries=3)
    reopen=r.fetch_query_results(query_id=REOPEN_QUERY_ID, max_retries=3)
    baselines=r.fetch_query_results(query_id=BASELINES_QUERY_ID, max_retries=3)
    return cfg, core, reopen, baselines

def baseline_summary(baselines):
    """Turn the [CS Perf] baselines rows into a compact, live summary for the LLM."""
    allrow=next((r for r in baselines if r['tag']=='__ALL__'), None)
    if not allrow or not allrow['n']:
        return {'avoidable_pct_of_all_reopens':None,'tag_bucket_leaders':[]}
    N=allrow['n']; base_av=allrow['avoidable_n']/N
    base_inc=allrow['incorrect_n']/N; base_incmp=allrow['incomplete_n']/N
    tags=[r for r in baselines if r['tag']!='__ALL__' and r['n']]
    def rec(r, lean):
        n=r['n']
        return {'tag':r['tag'],'n':n,'incorrect_pct':round(100*r['incorrect_n']/n,0),
                'incomplete_pct':round(100*r['incomplete_n']/n,0),
                'avoidable_lift':round((r['avoidable_n']/n)/base_av,2) if base_av else None,'leans':lean}
    inc_lead=sorted(tags,key=lambda r:(r['incorrect_n']/r['n'])/base_inc if base_inc else 0,reverse=True)[:8]
    incmp_lead=sorted(tags,key=lambda r:(r['incomplete_n']/r['n'])/base_incmp if base_incmp else 0,reverse=True)[:8]
    return {'avoidable_pct_of_all_reopens':round(100*base_av,1),
            'bucket_mix_pct':{'incorrect':round(100*base_inc,1),'incomplete':round(100*base_incmp,1)},
            'tag_bucket_leaders':[rec(r,'incorrect') for r in inc_lead]+[rec(r,'incomplete') for r in incmp_lead]}

# ==================== WEEK HELPERS ====================
def _week_label(week_start):
    d = week_start if isinstance(week_start, date) else pendulum.parse(str(week_start)).date()
    end = d + timedelta(days=6)
    return '%02d/%02d-%02d/%02d' % (d.day, d.month, end.day, end.month)

def _weeks_meta(core):
    """Return week_idx list oldest->newest and idx->label map from the core rows."""
    seen={}
    for r in core:
        seen[int(r['week_idx'])] = r['week_start']
    idxs=sorted(seen)                      # e.g. [1,2,3,4] (1=latest)
    old_to_new=list(reversed(idxs))        # [4,3,2,1]
    labels={i:_week_label(seen[i]) for i in idxs}
    return old_to_new, labels

# ==================== RENDER (v2 matrix) ====================
def _arrow(cur, prev, direction):
    if prev is None or cur is None: return ''
    d=cur-prev
    if abs(d)<1e-9: return f'<span style="font-size:9px;color:{GREY};margin-left:4px;">&#9644;</span>'
    up=d>0
    if direction=='neutral': col=GREY
    else:
        good=(d<0) if direction=='lower' else (d>0); col=UP if good else DOWN
    return f'<span style="font-size:9px;color:{col};margin-left:4px;">{"&#9650;" if up else "&#9660;"}</span>'

def _fmt_int(v):  return '&ndash;' if v is None else f'{int(round(v)):,}'
def _fmt_m(v):    return '&ndash;' if v is None else f'{v:g}m'
def _fmt_pct(v):  return '&ndash;' if v is None else f'{v:g}%'
def _fmt_plain(v):return '&ndash;' if v is None else f'{v:g}'

def _hdr(weeks):
    ths=[f'<th style="padding:7px 8px;border-bottom:2px solid {LINE};text-align:left;font-size:10px;letter-spacing:.4px;text-transform:uppercase;color:{FAINT};font-weight:600;">Metric</th>']
    for i,w in enumerate(weeks):
        newest=i==0; bg=f'background:{HL};' if newest else ''
        lbl=f'<div style="font-size:9px;color:{UP};font-weight:700;">LATEST</div>' if newest else ''
        ths.append(f'<th style="{bg}padding:7px 8px;border-bottom:2px solid {HLB if newest else LINE};text-align:right;font-size:10.5px;color:{MUT if newest else FAINT};font-weight:{"700" if newest else "600"};white-space:nowrap;">{lbl}{w}</th>')
    return f'<tr>{"".join(ths)}</tr>'

def _row(nw, label, primary, direction, pfmt=_fmt_plain, bracket=None, bfmt=_fmt_pct, cmp=None):
    """primary/bracket/cmp are oldest->newest lists (len nw). Displayed latest-left; arrow on latest only."""
    P=list(reversed(primary)); C=list(reversed(cmp if cmp is not None else primary))
    B=list(reversed(bracket)) if bracket is not None else None
    tds=[f'<td style="padding:7px 8px;border-bottom:1px solid {ROW};color:#334155;font-weight:500;">{label}</td>']
    for i in range(nw):
        newest=i==0; bg=f'background:{HL};' if newest else ''
        prev=C[i+1] if i+1<nw else None
        br=f' <span style="font-size:10px;color:{FAINT};">({bfmt(B[i])})</span>' if (B is not None and B[i] is not None) else ''
        st=f'{bg}padding:7px 8px;border-bottom:1px solid {ROW};text-align:right;white-space:nowrap;font-variant-numeric:tabular-nums;'
        st+=(f'font-weight:700;color:{INK};font-size:13.5px;' if newest else 'color:#475569;')
        ind=_arrow(C[i],prev,direction) if newest else ''
        tds.append(f'<td style="{st}">{pfmt(P[i])}{br}{ind}</td>')
    return f'<tr>{"".join(tds)}</tr>'

def _sep(nw): return f'<tr><td colspan="{nw+1}" style="padding:3px 0;"></td></tr>'
def _tbl(nw, weeks, rows): return f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;font-size:13px;">{_hdr(weeks)}{"".join(rows)}</table>'
def _sec(num,label,cap): return (f'<div style="font-size:11px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;color:{MUT};">{num} &middot; {label}</div><div style="font-size:12px;color:{FAINT};margin:2px 0 12px;">{cap}</div>')

def _series(by_idx, order, col):
    """by_idx: {week_idx: row}. order: oldest->newest week_idx list. Returns list of col values (or None)."""
    out=[]
    for i in order:
        r=by_idx.get(i)
        out.append(None if r is None or r.get(col) is None else r[col])
    return out

# ==================== ASSEMBLE (per agent) ====================
def assemble(core, reopen):
    order, labels = _weeks_meta(core)
    weeks_disp=[labels[i] for i in reversed(order)]  # latest-left
    nw=len(order)
    # index rows
    agents={}; shifts={}; teams={}
    for r in core:
        s=r['section']; wi=int(r['week_idx'])
        if s=='agent': agents.setdefault(r['agent_email'], {})[wi]=r
        elif s=='shift': shifts.setdefault((r['tier'],r['shift']), {})[wi]=r
        elif s=='team': teams.setdefault(r['tier'], {})[wi]=r
    reo_by_agent={}
    for r in reopen:
        reo_by_agent.setdefault(r['agent_email'], []).append(r)

    payloads=[]
    for email, aidx in agents.items():
        any_row=next(iter(aidx.values()))
        tier=any_row['tier']; shift=any_row['shift']; name=any_row['agent_name']
        tidx=teams.get(tier, {}); sidx=shifts.get((tier,shift), {})
        S=lambda idx,col: _series(idx, order, col)
        payloads.append(dict(
            email=email, name=name, tier=tier, shift=shift, nw=nw, weeks=weeks_disp,
            team={'total':S(tidx,'total'),'ow_n':S(tidx,'ow_n'),'pct_ow':S(tidx,'pct_ow'),'human_n':S(tidx,'human_n'),
                  'ow_p50':S(tidx,'ow_p50'),'hufrt_p50':S(tidx,'hufrt_p50'),'frt_p50':S(tidx,'frt_p50'),
                  'csat_n_ow':S(tidx,'csat_n_ow'),'csat_pos_ow':S(tidx,'csat_pos_ow'),'csat_n_hu':S(tidx,'csat_n_hu'),'csat_pos_hu':S(tidx,'csat_pos_hu'),
                  'reopen_n_ow':S(tidx,'reopen_n_ow'),'reopen_rate_ow':S(tidx,'reopen_rate_ow'),'reopen_n_hu':S(tidx,'reopen_n_hu'),'reopen_rate_hu':S(tidx,'reopen_rate_hu')},
            shift_m={'human_n':S(sidx,'human_n'),'hufrt_p50':S(sidx,'hufrt_p50'),'frt_p50':S(sidx,'frt_p50'),
                     'csat_n_hu':S(sidx,'csat_n_hu'),'csat_pos_hu':S(sidx,'csat_pos_hu'),'reopen_n_hu':S(sidx,'reopen_n_hu'),'reopen_rate_hu':S(sidx,'reopen_rate_hu')},
            you={'human_n':S(aidx,'human_n'),'hufrt_p50':S(aidx,'hufrt_p50'),'frt_p50':S(aidx,'frt_p50'),
                 'csat_n_hu':S(aidx,'csat_n_hu'),'csat_pos_hu':S(aidx,'csat_pos_hu'),'reopen_n_hu':S(aidx,'reopen_n_hu'),'reopen_rate_hu':S(aidx,'reopen_rate_hu')},
            buckets=_bucket_matrix(reo_by_agent.get(email, []), order),
            reopen_events=reo_by_agent.get(email, []),
        ))
    return payloads, order

def _bucket_matrix(events, order):
    """{bucket: [counts oldest->newest]} + avoidable/notfault/total series."""
    counts={b:{i:0 for i in order} for b in BUCKET_ORDER}
    for e in events:
        b=e.get('bucket'); wi=int(e['week_idx'])
        if b in counts and wi in counts[b]: counts[b][wi]+=1
    ser=lambda d:[d[i] for i in order]
    out={b:ser(counts[b]) for b in BUCKET_ORDER}
    out['avoidable']=[sum(out[b][k] for b in AVOID_BUCKETS) for k in range(len(order))]
    out['notfault']=[sum(out[b][k] for b in BUCKET_ORDER if b not in AVOID_BUCKETS) for k in range(len(order))]
    out['total']=[out['avoidable'][k]+out['notfault'][k] for k in range(len(order))]
    return out

# ==================== HTML EMAIL (per agent) ====================
def build_html(p, ai):
    nw=p['nw']; W=p['weeks']; T=p['team']; S=p['shift_m']; Y=p['you']; B=p['buckets']
    team=_tbl(nw,W,[
        _row(nw,'Total tickets',T['total'],'neutral',_fmt_int),
        _row(nw,'Overwatch tickets',T['ow_n'],'neutral',_fmt_int,bracket=T['pct_ow']),
        _row(nw,'Human tickets',T['human_n'],'neutral',_fmt_int),
        _sep(nw),
        _row(nw,'Created&rarr;OW (med)',T['ow_p50'],'lower',_fmt_m),
        _row(nw,'Escalated&rarr;human FRT',T['hufrt_p50'],'lower',_fmt_m),
        _row(nw,'Created&rarr;human FRT',T['frt_p50'],'lower',_fmt_m),
        _sep(nw),
        _row(nw,'OW CSAT (resp)',T['csat_n_ow'],'higher',_fmt_int,bracket=T['csat_pos_ow'],cmp=T['csat_pos_ow']),
        _row(nw,'Human CSAT (resp)',T['csat_n_hu'],'higher',_fmt_int,bracket=T['csat_pos_hu'],cmp=T['csat_pos_hu']),
        _row(nw,'OW reopen',T['reopen_n_ow'],'lower',_fmt_int,bracket=T['reopen_rate_ow'],cmp=T['reopen_rate_ow']),
        _row(nw,'Human reopen',T['reopen_n_hu'],'lower',_fmt_int,bracket=T['reopen_rate_hu'],cmp=T['reopen_rate_hu']),
    ])
    def human_tbl(D):
        return _tbl(nw,W,[
            _row(nw,'Human closes',D['human_n'],'neutral',_fmt_int),
            _row(nw,'Escalated&rarr;human FRT',D['hufrt_p50'],'lower',_fmt_m),
            _row(nw,'Created&rarr;human FRT',D['frt_p50'],'lower',_fmt_m),
            _row(nw,'Human CSAT (resp)',D['csat_n_hu'],'higher',_fmt_int,bracket=D['csat_pos_hu'],cmp=D['csat_pos_hu']),
            _row(nw,'Human reopen',D['reopen_n_hu'],'lower',_fmt_int,bracket=D['reopen_rate_hu'],cmp=D['reopen_rate_hu']),
        ])
    buckets=_tbl(nw,W,[
        _row(nw,'&#128308; Incorrect',B['incorrect'],'lower',_fmt_plain),
        _row(nw,'&#128992; Incomplete',B['incomplete'],'lower',_fmt_plain),
        _row(nw,'&nbsp;&nbsp;<b>Avoidable</b>',B['avoidable'],'lower',_fmt_plain),
        _sep(nw),
        _row(nw,'New issue',B['new_issue'],'neutral',_fmt_plain),
        _row(nw,'Clarification',B['clarification'],'neutral',_fmt_plain),
        _row(nw,'Noise',B['noise'],'neutral',_fmt_plain),
        _row(nw,'&nbsp;&nbsp;Not your fault',B['notfault'],'neutral',_fmt_plain),
        _sep(nw),
        _row(nw,'<b>Total reopen events</b>',B['total'],'lower',_fmt_plain),
    ])
    ab=lambda color,t,body:f'<div style="margin-bottom:14px;"><div style="font-size:12px;font-weight:700;color:{INK};margin-bottom:5px;"><span style="display:inline-block;width:7px;height:7px;border-radius:50%;background:{color};margin-right:7px;"></span>{t}</div>{body}</div>'
    ul=lambda items:'<ul style="margin:0;padding-left:18px;">'+''.join(f'<li style="font-size:12.5px;color:#334155;line-height:1.6;margin-bottom:3px;">{i}</li>' for i in items)+'</ul>'
    para=lambda t:f'<p style="font-size:12.5px;color:#334155;line-height:1.6;margin:0;">{t}</p>'
    dump_name=f"reopen_dump_{p['name'].split()[0].lower()}.xlsx"
    return f"""<!doctype html><html><body style="margin:0;padding:24px 0;background:#eef1f5;font-family:-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;">
<table role="presentation" align="center" width="640" cellpadding="0" cellspacing="0" style="width:640px;max-width:640px;margin:0 auto;background:#ffffff;border-radius:14px;overflow:hidden;">
  <tr><td style="background:#1e293b;padding:24px 26px;">
    <div style="font-size:19px;font-weight:650;color:#ffffff;">Your Weekly Performance Report</div>
    <div style="font-size:13px;color:#c7d2e0;margin-top:6px;line-height:1.5;">This is our understanding of your work this week, meant to help, not to grade. If any number looks off or you'd like something changed, just <b style="color:#fff;">reply to this email</b>.</div>
    <div style="margin-top:14px;font-size:12.5px;color:#e2e8f0;">
      <span style="background:rgba(255,255,255,.16);border-radius:6px;padding:2px 9px;margin-right:6px;">{p['tier']}</span>
      <span style="background:rgba(255,255,255,.16);border-radius:6px;padding:2px 9px;margin-right:8px;">Shift {p['shift']} IST</span>
      <b style="color:#fff;">{p['name']}</b></div>
  </td></tr>
  <tr><td style="padding:20px 26px;">{_sec('1','Team Performance',f"Your tier &middot; <b>{p['tier']}</b> &middot; closed-week &middot; assignee-credited &middot; arrow vs prior week")}{team}</td></tr>
  <tr><td style="padding:20px 26px;border-top:1px solid {LINE};">{_sec('2','Shift Performance',f"Your shift &middot; <b>{p['tier']} &middot; {p['shift']}</b> &middot; roster-aggregate &middot; human-only")}{human_tbl(S)}</td></tr>
  <tr><td style="padding:20px 26px;border-top:1px solid {LINE};">{_sec('3','Your Performance',f"{p['name']} &middot; human-only")}{human_tbl(Y)}
    <div style="margin-top:18px;">{_sec('','Reopen snapshots','per reopen event, credited to you as last-closer (+1 each time) &middot; bucket &times; week')}{buckets}</div>
    <div style="margin-top:12px;font-size:11.5px;color:{MUT};background:#f8fafc;border:1px dashed #dbe2ea;border-radius:8px;padding:9px 12px;">&#128206; <b>Reopen dump attached</b> (<i>{dump_name}</i>) &middot; every reopen grouped by bucket with ticket number and Trinity link.</div>
  </td></tr>
  <tr><td style="padding:20px 26px;border-top:1px solid {LINE};background:#f8fafc;">{_sec('4','AI Notes','Auto-generated from the tables above plus issue-type breakdown &middot; a read on the trend, not a verdict.')}
    {ab('#2a78d6','The trend', para(ai['trend']))}
    {ab('#15803d','What went well', ul(ai['strengths']))}
    {ab('#b91c1c','Where to tighten', ul(ai['weaknesses']))}
    {ab('#eda100','Suggested next steps', ul(ai['actions']))}
  </td></tr>
  <tr><td style="padding:18px 26px 26px;border-top:1px solid {LINE};text-align:center;font-size:12.5px;color:{MUT};line-height:1.6;">
    Just reply to this email with any questions, suggestions, or anything else.</td></tr>
</table></body></html>"""

# ==================== XLSX (per-agent reopen dump, latest week) ====================
def build_xlsx(p):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    latest=1  # week_idx 1 = latest week
    events=[e for e in p['reopen_events'] if int(e['week_idx'])==latest]
    events.sort(key=lambda e:(BUCKET_ORDER.index(e['bucket']) if e['bucket'] in BUCKET_ORDER else 9, e.get('reopen_ts','')))
    wb=Workbook(); ws=wb.active; ws.title='Reopen dump'
    thin=Side(style='thin',color='E2E6EC'); border=Border(bottom=thin)
    ws.column_dimensions['A'].width=16; ws.column_dimensions['B'].width=18; ws.column_dimensions['C'].width=62
    ws['A1']=f"Reopen dump - {p['name']} - {p['tier']} {p['shift']} - latest week - {len(events)} reopen events"
    ws['A1'].font=Font(bold=True,size=12); ws.merge_cells('A1:C1')
    HDR={'incorrect':'C23B22','incomplete':'E07D35','new_issue':'8A94A6','clarification':'9AA3AF','noise':'C2C8D0'}
    from collections import defaultdict
    by=defaultdict(list)
    for e in events: by[e['bucket']].append(e)
    r=3
    for b in BUCKET_ORDER:
        items=by.get(b,[])
        tag='AVOIDABLE' if b in AVOID_BUCKETS else 'not your fault'
        ws.cell(r,1,f'{BUCKET_NICE[b]}  ({len(items)})  - {tag}'); ws.cell(r,1).font=Font(bold=True,color='FFFFFF',size=11)
        for c in (1,2,3): ws.cell(r,c).fill=PatternFill('solid',fgColor=HDR[b])
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); r+=1
        ws.cell(r,1,'Ticket #'); ws.cell(r,2,'Reopened (IST)'); ws.cell(r,3,'Trinity link')
        for c in (1,2,3): ws.cell(r,c).font=Font(bold=True,size=9,color='64748B')
        r+=1
        for it in items:
            tn=ws.cell(r,1,f"#{it['ticket_number']}"); tn.hyperlink=it['trinity_link']; tn.font=Font(color='2A78D6',underline='single')
            ws.cell(r,2,it.get('reopen_ts','')).font=Font(color='475569',size=10)
            lk=ws.cell(r,3,it['trinity_link']); lk.hyperlink=it['trinity_link']; lk.font=Font(color='2A78D6',underline='single',size=10)
            for c in (1,2,3): ws.cell(r,c).border=border
            r+=1
        r+=1
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

# ==================== SECTION 4 - LLM (OpenAI-compatible REST) ====================
def _agent_top_tags(p):
    """This agent's latest-week avoidable reopens, by tag, split incorrect vs incomplete."""
    from collections import defaultdict
    d=defaultdict(lambda:{'incorrect':0,'incomplete':0})
    for e in p['reopen_events']:
        if int(e['week_idx'])==1 and e.get('bucket') in AVOID_BUCKETS and e.get('trinity_tags'):
            for t in str(e['trinity_tags']).split(', '):
                if t: d[t][e['bucket']]+=1
    out=[{'tag':t,'incorrect':v['incorrect'],'incomplete':v['incomplete'],'total':v['incorrect']+v['incomplete']}
         for t,v in d.items()]
    return sorted(out,key=lambda x:x['total'],reverse=True)[:8]

def ai_notes(p, cfg, baseline):
    def metric(series, unit=''):
        """Return a clearly-labeled {latest, prior, series_oldest_to_newest} block. series is oldest->newest."""
        latest=series[-1] if series else None
        prior=series[-2] if len(series)>1 else None
        return {'latest':(None if latest is None else f'{latest:g}{unit}'),
                'prior_week':(None if prior is None else f'{prior:g}{unit}'),
                'weekly_oldest_to_newest':[None if v is None else f'{v:g}{unit}' for v in series]}
    y=p['you']; s=p['shift_m']
    ctx={'agent':p['name'],'tier':p['tier'],'shift':p['shift'],'weeks_oldest_to_newest':list(p['weeks'][::-1]),
      'you':{
        'human_closes':metric(y['human_n']),
        'created_to_human_FRT_minutes':metric(y['frt_p50'],'m'),
        'escalated_to_human_FRT_minutes':metric(y['hufrt_p50'],'m'),
        'CSAT_percent_positive':metric(y['csat_pos_hu'],'%'),
        'CSAT_num_responses':metric(y['csat_n_hu']),
        'reopen_rate_percent':metric(y['reopen_rate_hu'],'%'),
        'reopen_count':metric(y['reopen_n_hu'])},
      'shift_median_for_context':{
        'created_to_human_FRT_minutes':metric(s['frt_p50'],'m'),
        'CSAT_percent_positive':metric(s['csat_pos_hu'],'%'),
        'reopen_rate_percent':metric(s['reopen_rate_hu'],'%')},
      'reopen_quality_latest_week':{
        'avoidable_total':p['buckets']['avoidable'][-1],
        'incorrect':p['buckets']['incorrect'][-1],'incomplete':p['buckets']['incomplete'][-1],
        'not_your_fault_total':p['buckets']['notfault'][-1]},
      'AGENT_TOP_TAGS':_agent_top_tags(p),
      'POPULATION_BASELINE':baseline,
      'TAG_BUCKET_LEADERS':baseline.get('tag_bucket_leaders', [])}
    # The analysis framework (the "skill") lives in the config query column ai_rubric - editable in
    # Redash, no code push. Fall back to a minimal instruction if the column is missing.
    rubric=(cfg.get('ai_rubric') or '').strip() or (
      "Write supportive weekly AI notes. Use only the numbers in DATA, quoting them exactly. "
      "Return STRICT JSON: trend (string), strengths (2), weaknesses (2), actions (2).")
    prompt=rubric+"\n\nDATA:\n"+json.dumps(ctx, default=str)
    try:
        body=json.dumps({'model':cfg['llm_model'],'max_tokens':600,'temperature':0.4,
            'messages':[{'role':'user','content':prompt}]}).encode()
        req=urllib.request.Request(cfg['llm_proxy_url'].rstrip('/')+'/chat/completions', data=body,
            headers={'Authorization':'Bearer '+cfg['llm_proxy_api_key'],'Content-Type':'application/json'})
        resp=json.loads(urllib.request.urlopen(req, timeout=60).read())
        txt=resp['choices'][0]['message']['content'].strip()
        if txt.startswith('```'): txt=txt.strip('`').split('\n',1)[1] if '\n' in txt else txt.strip('`')
        j=json.loads(txt)
        return {'trend':j['trend'],'strengths':j['strengths'][:3],'weaknesses':j['weaknesses'][:3],'actions':j['actions'][:3]}
    except Exception as e:
        logger.warning('AI notes fell back for %s: %s', p['email'], e)
        return _ai_fallback(p)

def _ai_fallback(p):
    y=p['you']; frt=y['frt_p50']; reop=y['reopen_rate_hu']
    def dlt(a):
        if len(a)<2 or a[-1] is None or a[-2] is None: return 'held steady'
        return 'improved' if a[-1]<a[-2] else ('rose' if a[-1]>a[-2] else 'held steady')
    return {
      'trend': f"This week you closed {y['human_n'][-1]} tickets; FRT {dlt(frt)} and reopen rate {dlt(reop)} versus last week.",
      'strengths':['Consistent weekly volume.','Metrics tracked and shared for transparency.'],
      'weaknesses':['Review the avoidable reopen buckets (incorrect / incomplete) in the attached dump.'],
      'actions':['Confirm the fix and restate the customer request before closing.','Check the reopen dump to spot recurring issue types.'],
    }

# ==================== DELIVERY ====================
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT   = 587
GMAIL_SECRET_NAME = 'cs-perf-gmail-app-password'   # 16-char Gmail App Password for from_email

def _gmail_app_password():
    """App password: env override for local tests, else Secret Manager (Composer)."""
    pw = os.getenv('CS_PERF_GMAIL_APP_PASSWORD')
    if pw:
        return pw
    from utils.secrets import get_secret
    return get_secret(GMAIL_SECRET_NAME)

def _plain_fallback():
    return ("Your weekly performance report is in the HTML body of this email, with your reopen "
            "dump attached as an .xlsx. Questions, suggestions, or anything else - just reply.")

def send_email(to, subject, html, xlsx_bytes, filename, cfg):
    """Send one agent's report as HTML + .xlsx attachment via Gmail SMTP (app password).
    DRY_RUN writes HTML+xlsx to disk instead of sending."""
    if DRY_RUN:
        os.makedirs(DRY_RUN_DIR, exist_ok=True)
        base=os.path.join(DRY_RUN_DIR, to.split('@')[0])
        open(base+'.html','w').write(html); open(base+'_'+filename,'wb').write(xlsx_bytes)
        logger.info('[DRY_RUN] wrote %s.html + %s (%d bytes xlsx)', base, filename, len(xlsx_bytes))
        return
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    sender = cfg.get('from_email'); reply_to = cfg.get('reply_to') or sender
    if not sender:
        raise RuntimeError('from_email missing in config query')
    msg = MIMEMultipart('mixed')
    msg['Subject'] = subject; msg['From'] = 'CS Weekly Report <%s>' % sender
    msg['To'] = to; msg['Reply-To'] = reply_to
    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(_plain_fallback(), 'plain'))
    alt.attach(MIMEText(html, 'html'))
    msg.attach(alt)
    att = MIMEApplication(xlsx_bytes, _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    att.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(att)
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=60) as s:
        s.starttls(); s.login(sender, _gmail_app_password())
        s.sendmail(sender, [to], msg.as_string())
    logger.info('sent report to %s (xlsx %d bytes)', to, len(xlsx_bytes))

# ==================== MAIN TASK ====================
def run_perf_report(**context):
    cfg, core, reopen, baselines = fetch_all()
    payloads, _ = assemble(core, reopen)
    base = baseline_summary(baselines)
    logger.info('assembled %d agent reports (%d core, %d reopen rows); baseline avoidable=%s%%',
                len(payloads), len(core), len(reopen), base.get('avoidable_pct_of_all_reopens'))
    subject_prefix = cfg.get('subject_prefix') or 'Your Weekly Performance Report'
    test_to = os.getenv('CS_PERF_TEST_RECIPIENT')             # pilot: route ALL emails here
    limit   = int(os.getenv('CS_PERF_LIMIT', '0') or 0)       # pilot: only first N agents (0 = all)
    if limit: payloads = payloads[:limit]
    if test_to: logger.warning('PILOT: routing all %d emails to %s', len(payloads), test_to)
    sent=0; failed=[]
    for p in payloads:
        try:
            ai = ai_notes(p, cfg, base)                       # LLM (has its own fallback)
            html = build_html(p, ai)
            xlsx = build_xlsx(p)
            fname=f"reopen_dump_{p['name'].split()[0].lower().replace('/','-') or 'agent'}.xlsx"
            to = test_to or p['email']
            subj = ('[TEST %s] ' % p['name'] + subject_prefix) if test_to else subject_prefix
            send_email(to, subj, html, xlsx, fname, cfg)
            sent+=1
        except Exception as e:                                # one bad agent must not sink the run
            logger.exception('perf report FAILED for %s: %s', p.get('email'), e)
            failed.append(p.get('email'))
    logger.info('CS WEEKLY PERFORMANCE REPORT: %s for %d/%d agents (%d failed: %s)',
                'DRY_RUN wrote' if DRY_RUN else 'sent', sent, len(payloads), len(failed), failed)
    if failed and not DRY_RUN:
        logger.warning('agents skipped due to errors: %s', failed)

# ==================== DAG ====================
default_args = {
    'owner': 'cs_team', 'depends_on_past': False,
    'start_date': pendulum.datetime(2025, 1, 1, tz='Asia/Kolkata'),
    'email_on_failure': False, 'email_on_retry': False,
    'retries': 1, 'retry_delay': timedelta(minutes=3),
}
dag = DAG(
    'cs_performance_report_weekly',
    default_args=default_args,
    description='Per-agent weekly performance email (v2 matrix); numbers in Redash 41791/41792, config 41839',
    schedule_interval='0 19 * * 1',   # Monday 19:00 IST
    catchup=False, is_paused_upon_creation=True,
    tags=['email', 'trinity', 'performance', 'cs_reports', 'cs_team'],
)
PythonOperator(task_id='run_perf_report', python_callable=run_perf_report, dag=dag)
